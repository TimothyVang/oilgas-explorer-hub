import argparse
import html
import json
import shutil
import subprocess
import zipfile
from pathlib import Path
from urllib.parse import quote
from xml.etree import ElementTree

from openpyxl import load_workbook


ZIP_ROOT = "BAH videos and operational ledger"
PANDOC = "pandoc"
CHROME_CANDIDATES = [
    Path("C:/Program Files/Google/Chrome/Application/chrome.exe"),
    Path("C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"),
    Path("C:/Program Files/Microsoft/Edge/Application/msedge.exe"),
    Path("C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe"),
]
FFMPEG_CANDIDATES = [
    Path("C:/ffmpeg/ffmpeg-8.0.1-essentials_build/bin/ffmpeg.exe"),
]

ASSETS = [
    {
        "source": "cv resume/Bryant Mook CV_8_2_2024_r1.1.pdf",
        "title": "Technical Leadership CV",
        "description": "Background packet for technical leadership credentials and investor review.",
        "category": "management",
        "asset_type": "document",
        "sort_order": 1,
        "is_featured": False,
        "preview_name": "technical-leadership-cv.pdf",
    },
    {
        "source": "cv resume/Bryant Mook CV_8_2_2024_r1.1.docx",
        "title": "Technical Leadership CV Source Preview",
        "description": "Browser preview of the editable technical leadership background document.",
        "category": "management",
        "asset_type": "document",
        "sort_order": 2,
        "is_featured": False,
        "preview_name": "technical-leadership-cv-source-preview.pdf",
    },
    {
        "source": "Investor Pitch Deck/1. BAH Oil - Jackson Sands Teaser.docx",
        "title": "Start Here: Investor Teaser",
        "description": "Read this first for the concise Jackson Sands opportunity overview.",
        "category": "overview",
        "asset_type": "document",
        "sort_order": 1,
        "is_featured": True,
        "preview_name": "investor-teaser-preview.pdf",
    },
    {
        "source": "Investor Pitch Deck/2. BAH Oil - Jackson Sands Deal Snapshot.docx",
        "title": "Deal Snapshot",
        "description": "Quick deal snapshot covering the review sequence and key investor materials.",
        "category": "pitch",
        "asset_type": "document",
        "sort_order": 2,
        "is_featured": False,
        "preview_name": "deal-snapshot-preview.pdf",
    },
    {
        "source": "Investor Pitch Deck/3. BAH Oil - Jackson Sands Pitch Deck.pptx",
        "title": "Investor Presentation Deck",
        "description": "Slide preview for the Jackson Sands investor presentation deck.",
        "category": "pitch",
        "asset_type": "document",
        "sort_order": 3,
        "is_featured": False,
        "preview_name": "investor-presentation-deck-preview.pdf",
    },
    {
        "source": "Investor Pitch Deck/4. BAH Oil - Jackson Sands Pitch Deck (Word version).docx",
        "title": "Investor Overview Memorandum",
        "description": "Long-form browser preview supporting the investor presentation deck.",
        "category": "pitch",
        "asset_type": "document",
        "sort_order": 4,
        "is_featured": False,
        "preview_name": "investor-overview-memorandum-preview.pdf",
    },
    {
        "source": "Investor Pitch Deck/5. AFE - Investor Print 4.24.26.xlsx",
        "title": "AFE Budget Preview",
        "description": "Review the budget and AFE support in a browser-safe PDF preview.",
        "category": "financials",
        "asset_type": "document",
        "sort_order": 1,
        "is_featured": False,
        "preview_name": "afe-budget-preview.pdf",
    },
    {
        "source": "Investor Pitch Deck/6. Pro Forma - Investor Locked Cells.xlsx",
        "title": "Pro Forma Model Preview",
        "description": "Browser-safe preview of the investor pro forma and economics model.",
        "category": "financials",
        "asset_type": "document",
        "sort_order": 2,
        "is_featured": False,
        "preview_name": "pro-forma-model-preview.pdf",
    },
    {
        "source": "Investor Pitch Deck/7. Charco_Redondo_Jackson_Sands - 3_well.PDF",
        "title": "Technical Economics Appendix",
        "description": "Technical economics appendix for the Jackson Sands three-well case.",
        "category": "financials",
        "asset_type": "document",
        "sort_order": 3,
        "is_featured": False,
        "preview_name": "technical-economics-appendix.pdf",
    },
    {
        "source": "Investor Pitch Deck/8. Economic_Summary_Charco_Redondo_Jackson_Sands - 3_well.PDF",
        "title": "Economic Summary",
        "description": "One-page economics summary for fast investor review.",
        "category": "financials",
        "asset_type": "document",
        "sort_order": 4,
        "is_featured": False,
        "preview_name": "economic-summary.pdf",
    },
    {
        "source": "Investor Pitch Deck/9. Jackson Mapping Deck.pdf",
        "title": "Jackson Mapping Deck",
        "description": "Review maps and technical location support after the overview materials.",
        "category": "mapping",
        "asset_type": "document",
        "sort_order": 1,
        "is_featured": False,
        "preview_name": "jackson-mapping-deck.pdf",
    },
    {
        "source": "ops ledger/IMG_5121.heic",
        "title": "Operations Reference Image",
        "description": "Browser preview image supporting the operations review materials.",
        "category": "operations",
        "asset_type": "image",
        "sort_order": 6,
        "is_featured": False,
        "preview_name": "operations-reference-image.jpg",
    },
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("zip_path", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()

    chrome = find_first(CHROME_CANDIDATES, "Chrome or Edge")
    ffmpeg = find_first(FFMPEG_CANDIDATES, "ffmpeg", required=False)
    extracted = args.output_dir / "extracted"
    previews = args.output_dir / "previews"
    work = args.output_dir / "work"
    for path in [extracted, previews, work]:
        path.mkdir(parents=True, exist_ok=True)

    extract_zip(args.zip_path, extracted)

    manifest = []
    for asset in ASSETS:
        source = find_source(extracted, asset["source"])
        preview_path = previews / asset["preview_name"]
        suffix = source.suffix.lower()
        if suffix == ".pdf":
            shutil.copyfile(source, preview_path)
        elif suffix == ".docx":
            docx_to_pdf(source, preview_path, work, chrome)
        elif suffix == ".xlsx":
            xlsx_to_pdf(source, preview_path, work, chrome)
        elif suffix == ".pptx":
            pptx_to_pdf(source, preview_path, work, chrome)
        elif suffix == ".heic":
            if ffmpeg is None:
                raise RuntimeError("ffmpeg is required to convert HEIC previews")
            heic_to_jpg(source, preview_path, ffmpeg)
        else:
            raise RuntimeError(f"Unsupported source type: {source}")

        storage_dir = "images" if asset["asset_type"] == "image" else "documents"
        mime_type = "image/jpeg" if asset["asset_type"] == "image" else "application/pdf"
        manifest.append({
            "title": asset["title"],
            "description": asset["description"],
            "category": asset["category"],
            "asset_type": asset["asset_type"],
            "sort_order": asset["sort_order"],
            "is_featured": asset["is_featured"],
            "original_filename": source.name,
            "source_size": source.stat().st_size,
            "preview_path": str(preview_path),
            "preview_size": preview_path.stat().st_size,
            "mime_type": mime_type,
            "storage_path": f"briefing-20260524/{storage_dir}/{asset['preview_name']}",
        })

    manifest_path = args.output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(manifest_path)


def find_first(paths: list[Path], label: str, required: bool = True) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    if required:
        raise RuntimeError(f"Could not find {label}")
    return None


def extract_zip(zip_path: Path, dest: Path) -> None:
    marker = dest / ".extracted"
    if marker.exists():
        return
    with zipfile.ZipFile(zip_path) as archive:
        archive.extractall(dest)
    marker.write_text("ok", encoding="utf-8")


def find_source(extracted: Path, relative: str) -> Path:
    target = normalize_name((ZIP_ROOT + "/" + relative).replace("\\", "/"))
    matches = [path for path in extracted.rglob("*") if normalize_name(path.as_posix()).endswith(target)]
    if len(matches) != 1:
        raise RuntimeError(f"Expected one match for {relative}, found {len(matches)}")
    return matches[0]


def normalize_name(value: str) -> str:
    return value.replace("�", "-").replace("–", "-").replace("—", "-")


def docx_to_pdf(source: Path, output: Path, work: Path, chrome: Path) -> None:
    html_path = work / f"{source.stem}.html"
    subprocess.run([
        PANDOC,
        str(source),
        "--standalone",
        "--embed-resources",
        "--metadata",
        f"title={source.stem}",
        "-o",
        str(html_path),
    ], check=True)
    html_to_pdf(html_path, output, chrome)


def xlsx_to_pdf(source: Path, output: Path, work: Path, chrome: Path) -> None:
    workbook = load_workbook(source, data_only=True)
    sections = []
    for sheet in workbook.worksheets:
        rows = []
        max_row = min(sheet.max_row, 80)
        max_col = min(sheet.max_column, 14)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            cells = "".join(f"<td>{html.escape(format_cell(cell))}</td>" for cell in row)
            rows.append(f"<tr>{cells}</tr>")
        sections.append(f"<section><h2>{html.escape(sheet.title)}</h2><table>{''.join(rows)}</table></section>")
    html_path = work / f"{source.stem}.html"
    html_path.write_text(wrap_html(source.stem, "\n".join(sections), landscape=True), encoding="utf-8")
    html_to_pdf(html_path, output, chrome)


def pptx_to_pdf(source: Path, output: Path, work: Path, chrome: Path) -> None:
    slides = []
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    with zipfile.ZipFile(source) as archive:
        slide_names = sorted(
            [name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")],
            key=lambda value: int("".join(ch for ch in Path(value).stem if ch.isdigit()) or "0"),
        )
        for index, name in enumerate(slide_names, start=1):
            root = ElementTree.fromstring(archive.read(name))
            texts = [node.text.strip() for node in root.findall(".//a:t", ns) if node.text and node.text.strip()]
            body = "".join(f"<p>{html.escape(text)}</p>" for text in texts) or "<p>No extractable slide text.</p>"
            slides.append(f"<section class='slide'><h2>Slide {index}</h2>{body}</section>")
    html_path = work / f"{source.stem}.html"
    html_path.write_text(wrap_html(source.stem, "\n".join(slides)), encoding="utf-8")
    html_to_pdf(html_path, output, chrome)


def heic_to_jpg(source: Path, output: Path, ffmpeg: Path) -> None:
    subprocess.run([str(ffmpeg), "-y", "-i", str(source), "-frames:v", "1", str(output)], check=True)


def html_to_pdf(html_path: Path, output: Path, chrome: Path) -> None:
    url = html_path.resolve().as_uri()
    subprocess.run([
        str(chrome),
        "--headless=new",
        "--disable-gpu",
        "--no-sandbox",
        f"--print-to-pdf={output}",
        url,
    ], check=True)


def wrap_html(title: str, body: str, landscape: bool = False) -> str:
    page = "landscape" if landscape else "portrait"
    return f"""<!doctype html>
<html>
<head>
  <meta charset=\"utf-8\" />
  <title>{html.escape(title)}</title>
  <style>
    @page {{ size: Letter {page}; margin: 0.45in; }}
    body {{ font-family: Arial, sans-serif; color: #111827; }}
    h1 {{ font-size: 24px; margin: 0 0 18px; }}
    h2 {{ font-size: 18px; margin: 24px 0 10px; color: #0f3554; }}
    p {{ font-size: 12px; line-height: 1.45; margin: 6px 0; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 9px; page-break-inside: auto; }}
    tr {{ page-break-inside: avoid; }}
    td {{ border: 1px solid #d1d5db; padding: 4px; vertical-align: top; }}
    .slide {{ page-break-after: always; border: 1px solid #d1d5db; padding: 18px; min-height: 6.8in; }}
  </style>
</head>
<body>
  <h1>{html.escape(title)}</h1>
  {body}
</body>
</html>"""


def format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value)


if __name__ == "__main__":
    main()
